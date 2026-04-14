import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# -----------------------------
# Load Excel Data
# -----------------------------
file_path = "marks.xlsx"
df = pd.read_excel(file_path)

# -----------------------------
# Calculate Total & Percentage
# -----------------------------
df["Total"] = df[["Math", "Science", "English"]].sum(axis=1)
df["Percentage"] = df["Total"] / 3

# -----------------------------
# Assign Grades
# -----------------------------
def assign_grade(p):
    if p >= 90:
        return "A"
    elif p >= 75:
        return "B"
    elif p >= 60:
        return "C"
    else:
        return "D"

df["Grade"] = df["Percentage"].apply(assign_grade)

# -----------------------------
# Sort Data (Highest First)
# -----------------------------
df = df.sort_values(by="Total", ascending=False)

# -----------------------------
# Save to Excel
# -----------------------------
output_file = "student_report.xlsx"
df.to_excel(output_file, index=False)

# -----------------------------
# Apply Conditional Formatting
# -----------------------------
wb = load_workbook(output_file)
ws = wb.active

# Colors
green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
orange = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
red = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

# Apply color based on grade
for row in range(2, ws.max_row + 1):
    grade_cell = ws[f"G{row}"]  # Assuming Grade column is G

    if grade_cell.value == "A":
        grade_cell.fill = green
    elif grade_cell.value == "B":
        grade_cell.fill = yellow
    elif grade_cell.value == "C":
        grade_cell.fill = orange
    else:
        grade_cell.fill = red

wb.save(output_file)

print("✅ Student report generated with formatting!")